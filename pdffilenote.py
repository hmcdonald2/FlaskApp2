import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import os
import docx
import os.path
# import PyPDF2
from fpdf import FPDF

def testy73():
    try:
        m876 = request.form.get("mname")
        print(m876)
    except:
        print("testy73 function failed")

def testing79():


def mkpdffilenote():
    from fpdf import FPDF
    import os
    import os.path
    try:
        # obtain matter name
        m = request.form.get("mname")
        mname = m.upper()
        # obtain date
        date = now
    except:
        print("failed to set matter name and date")
    try:
        # obtain attendance type
        xax = ""
        for el in attendtype:
            if request.form.get(el) == "on":
                xax += el + ", "
        tp = xax
        type = tp.upper()
    except:
        print("failed to set attend type")
    try:
        # obtain attendance with
        xab = ""
        for el in attendupon:
            if request.form.get(el) == "on":
                xab += el + ", "
        attendon = xab.upper()
        # obtain text of file note
    except:
        print("failed to set attendance with")
    try:
        notes = request.form.get("text")
        # create the pdf using the now set variables
    except:
        print("failed to set notes for pdf")
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(True, 2)
        pdf.set_font("Arial", size = 20)
        pdf.cell(200,10, txt = "FILE NOTE", ln = 1, align = 'C')
        pdf.set_font("Arial", size = 12)
        pdf.cell(200,10, txt = f"Matter name:    {mname}", ln = 1, align = 'L')
        pdf.cell(200,10, txt = f"Date:             {date}", ln = 1, align = 'l')
        pdf.cell(200,10, txt = f"Attendance Type: {type}", ln = 1, align = 'l')
        pdf.cell(200,10, txt = f"Attendance Upon: {attendon}, ln = 1, align = 'l")
        pdf.cell(200, 10, txt="Author: Harry McDonald", ln=1)
        pdf.cell(200,10, txt = "NOTES:", ln = 1)
        pdf.multi_cell(200,10, txt = notes, align = 'l')
        pdf.ln()
        # save the pdf under the right file name
        for k in range(1,400):
            if not os.path.exists(f"File Note {mname} {date}.pdf"):
                pdf.output(f'File Note {mname} {date}.pdf')
            elif not os.path.exists(f"File Note {mname} {date}({k}).pdf"):
                pdf.output(f"File Note {mname} {date}({k}).pdf")
                break
        # todo use os.mkdir or shutil move to get file notes into folders sorted by matter

        # if you want to save as word documents instead
        # d = docx.Document()
        # d.add_paragraph(f"FILE NOTE\n{mname}\n{date}\n{type}\n{notes}")
        # for k in range(1,400):
        #     if not os.path.exists(f"Test{mname}.docx"):
        #         d.save(f'Test{mname}.docx')
        #     elif not os.path.exists(f"Test{mname}{k}.docx"):
        #         d.save(f'Test{mname}{k}.docx')
        #         break
    except:
        print(f"Failed to save the file note as a PDF")

